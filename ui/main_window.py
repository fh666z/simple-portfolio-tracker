"""Main window for Portfolio Tracker."""
import csv
import json
from datetime import datetime
from pathlib import Path
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QTabWidget, QPushButton,
    QHBoxLayout, QMessageBox, QStatusBar, QFileDialog, QInputDialog,
    QProgressDialog, QApplication, QMenu, QDialog, QDialogButtonBox,
    QTextBrowser, QLabel
)
from PyQt6.QtCore import Qt, QTimer, QUrl
from PyQt6.QtGui import QShortcut, QKeySequence, QAction, QDesktopServices

from core.models import Portfolio, Holding
from core.calculator import PortfolioCalculator
from core.data_parser import parse_file
from core.ocr_parser import parse_image_file, check_tesseract
from core.persistence import MappingsStore, SettingsStore, PortfolioStore, get_data_dir
from core import __version__

from .portfolio_tab import PortfolioTab
from .instrument_config_tab import InstrumentConfigTab
from .stats_tab import StatsTab
from .currency_tab import CurrencyTab
from .import_dialog import ImportDialog
from .review_dialog import ReviewDialog


_USER_GUIDE_TEXT = """Portfolio Tracker – User Guide

Menu & shortcuts
• File: Import Portfolio Data, Load Data, Export (JSON / CSV / Excel), Reset Data, Exit.
• Edit: Find (focus search in Portfolio tab).
• Help: User Guide, Keyboard Shortcuts, Data Storage, About.
• Shortcuts: Ctrl+N Import, Ctrl+O Load, Ctrl+S Export, Ctrl+F Find, Ctrl+Q Exit.

Import & data
• Import from images (PNG, JPG – OCR via Tesseract) or spreadsheets (XLSX, XLS, CSV).
• Drag & drop or Browse in the import dialog. Review dialog lets you edit extracted data and fix OCR errors.
• Load Data: Open a previously exported JSON file (restores portfolio, currencies, rates, mappings).
• Reset Data: Clear all holdings and free cash; requires typing "DELETE" to confirm.

Export
• Export to JSON (Ctrl+S): Full snapshot (portfolio + settings + mappings). Default filename includes date.
• Export to CSV or Excel: Holdings table for spreadsheets or formatted XLSX.

Portfolio tab
• Holdings table: Instrument, Position, Last Price, Market Value, Value (EUR), Cost Basis, Allocation %, Target %, Diff, Unrealized P&L.
• Editable: Position, Last Price, Cost Basis, Target %, Unrealized P&L; Type and Region via dropdowns.
• Delete: × button per row or right-click → Delete.
• Search and Type filter (Equity, Bonds, etc.). Summary: Total Invested, Free Cash, Total (EUR).

Instrument Config tab
• Set Currency, Type (Equity, Bonds, Commodity, etc.), Region (US, EU, EM, etc.) per instrument. Sortable; changes saved.

Currency Exchange tab
• Edit rate (units per 1 EUR). Update rates from internet (Frankfurter API, no key). Add/remove currencies; EUR fixed at 1.0.

Statistics tab
• Pie charts: By Type, By Region, Detailed (Type + Region). Charts update when portfolio or config changes.

Persistence
• Data is saved automatically. Stored in your user data directory (see Help → Data Storage).
• Window size, tab order, and table column order are restored on restart.

Prerequisites
• Python 3.10+. Tesseract OCR optional for image import (see README or About for install link).
"""


class MainWindow(QMainWindow):
    """Main application window."""
    
    def __init__(self):
        super().__init__()
        
        # Initialize stores
        self.mappings_store = MappingsStore()
        self.settings_store = SettingsStore()
        self.portfolio_store = PortfolioStore()
        
        # Initialize calculator with empty or loaded portfolio
        portfolio = self.portfolio_store.load() or Portfolio()
        self.calculator = PortfolioCalculator(portfolio, self.settings_store)
        
        # Apply saved mappings to loaded holdings
        self.mappings_store.apply_mappings(self.calculator.portfolio.holdings)
        
        # Set free cash from settings
        saved_free_cash = self.settings_store.get('free_cash', 0)
        self.calculator.set_free_cash(saved_free_cash)
        
        self.setup_ui()
        self.refresh_all()
    
    def setup_ui(self):
        """Set up the main window UI."""
        self.setWindowTitle("Portfolio Tracker")
        self.setMinimumSize(1100, 650)
        
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout(central_widget)
        
        self._setup_menus()
        
        # Tab widget
        self.tabs = QTabWidget()
        self.tabs.setMovable(True)  # Enable tab reordering via drag-and-drop
        
        # Tab names for persistence (used as identifiers)
        self.tab_names = []
        
        # Portfolio tab
        self.portfolio_tab = PortfolioTab(self.calculator, self.settings_store)
        self.portfolio_tab.portfolio_changed.connect(self.on_portfolio_changed)
        self.portfolio_tab.import_requested.connect(self.on_new_input)
        self.tabs.addTab(self.portfolio_tab, "Portfolio")
        self.tab_names.append("portfolio")
        
        # Instrument Config tab
        self.config_tab = InstrumentConfigTab(self.calculator, self.settings_store)
        self.config_tab.config_changed.connect(self.on_config_changed)
        self.tabs.addTab(self.config_tab, "Instrument Config")
        self.tab_names.append("instrument_config")
        
        # Stats tab
        self.stats_tab = StatsTab(self.calculator, self.settings_store)
        self.tabs.addTab(self.stats_tab, "Statistics")
        self.tab_names.append("statistics")
        
        # Currency Exchange tab
        self.currency_tab = CurrencyTab(self.settings_store)
        self.currency_tab.rates_changed.connect(self.on_rates_changed)
        self.tabs.addTab(self.currency_tab, "Currency Exchange")
        self.tab_names.append("currency_exchange")
        
        # Connect tab moved signal for persistence
        self.tabs.tabBar().tabMoved.connect(self.on_tab_moved)
        
        # Restore saved tab order
        self.restore_tab_order()
        
        layout.addWidget(self.tabs)
        
        # Status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.update_status_bar()
        
        # Restore window geometry
        geometry = self.settings_store.get('window_geometry')
        if geometry:
            try:
                self.restoreGeometry(bytes.fromhex(geometry))
            except Exception:
                pass
        
        # Set up keyboard shortcuts
        self.setup_shortcuts()
    
    def _setup_menus(self):
        """Set up the application menu bar (File, Edit, Help)."""
        menubar = self.menuBar()
        
        # File menu
        file_menu = menubar.addMenu("&File")
        
        import_action = QAction("Import Portfolio Data...", self)
        import_action.setShortcut(QKeySequence("Ctrl+N"))
        import_action.setStatusTip("Import portfolio data from image or spreadsheet")
        import_action.triggered.connect(self.on_new_input)
        file_menu.addAction(import_action)
        
        load_action = QAction("Load Data...", self)
        load_action.setShortcut(QKeySequence("Ctrl+O"))
        load_action.setStatusTip("Load portfolio from a JSON file")
        load_action.triggered.connect(self.on_load_data)
        file_menu.addAction(load_action)
        
        file_menu.addSeparator()
        
        # Export submenu
        export_submenu = QMenu("Export", self)
        export_json_action = QAction("Export to JSON (Ctrl+S)", self)
        export_json_action.setShortcut(QKeySequence("Ctrl+S"))
        export_json_action.triggered.connect(self.on_save_data)
        export_submenu.addAction(export_json_action)
        export_csv_action = QAction("Export to CSV", self)
        export_csv_action.triggered.connect(self.on_export_csv)
        export_submenu.addAction(export_csv_action)
        export_excel_action = QAction("Export to Excel", self)
        export_excel_action.triggered.connect(self.on_export_excel)
        export_submenu.addAction(export_excel_action)
        file_menu.addMenu(export_submenu)
        
        file_menu.addSeparator()
        
        reset_action = QAction("Reset Data...", self)
        reset_action.setStatusTip("Clear all portfolio data (requires confirmation)")
        reset_action.triggered.connect(self.on_reset_data)
        file_menu.addAction(reset_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction("E&xit", self)
        exit_action.setShortcut(QKeySequence("Ctrl+Q"))
        exit_action.setStatusTip("Quit the application")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Edit menu
        edit_menu = menubar.addMenu("&Edit")
        find_action = QAction("Find...", self)
        find_action.setShortcut(QKeySequence("Ctrl+F"))
        find_action.setStatusTip("Focus search in Portfolio tab")
        find_action.triggered.connect(self.focus_search)
        edit_menu.addAction(find_action)
        
        # Help menu
        help_menu = menubar.addMenu("&Help")
        user_guide_action = QAction("User Guide", self)
        user_guide_action.setStatusTip("Open the user guide")
        user_guide_action.triggered.connect(self.on_user_guide)
        help_menu.addAction(user_guide_action)
        shortcuts_action = QAction("Keyboard Shortcuts", self)
        shortcuts_action.setStatusTip("Show keyboard shortcuts")
        shortcuts_action.triggered.connect(self.on_shortcuts)
        help_menu.addAction(shortcuts_action)
        data_storage_action = QAction("Data Storage...", self)
        data_storage_action.setStatusTip("Show where data is stored")
        data_storage_action.triggered.connect(self.on_data_storage)
        help_menu.addAction(data_storage_action)
        help_menu.addSeparator()
        about_action = QAction("About", self)
        about_action.setStatusTip("About Portfolio Tracker")
        about_action.triggered.connect(self.on_about)
        help_menu.addAction(about_action)

    def on_user_guide(self):
        """Show the User Guide dialog."""
        dialog = QDialog(self)
        dialog.setWindowTitle("User Guide")
        dialog.setMinimumSize(520, 480)
        layout = QVBoxLayout(dialog)
        browser = QTextBrowser(dialog)
        browser.setOpenExternalLinks(True)
        browser.setPlainText(_USER_GUIDE_TEXT)
        layout.addWidget(browser)
        close_btn = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        close_btn.rejected.connect(dialog.close)
        layout.addWidget(close_btn)
        dialog.exec()

    def on_shortcuts(self):
        """Show the Keyboard Shortcuts dialog."""
        QMessageBox.information(
            self,
            "Keyboard Shortcuts",
            "Ctrl+N – Import portfolio data\n"
            "Ctrl+O – Load data (JSON)\n"
            "Ctrl+S – Export to JSON\n"
            "Ctrl+F – Focus search (Portfolio tab)\n"
            "Ctrl+Q – Exit"
        )

    def on_data_storage(self):
        """Show the Data Storage Location dialog."""
        data_dir = get_data_dir()
        path_str = str(data_dir.resolve())
        files_note = "Files: portfolio.json, mappings.json, settings.json"
        dialog = QDialog(self)
        dialog.setWindowTitle("Data Storage Location")
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel("Your portfolio data is stored in:"))
        path_label = QLabel(path_str)
        path_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        path_label.setWordWrap(True)
        layout.addWidget(path_label)
        layout.addWidget(QLabel(files_note))
        bbox = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        open_btn = bbox.addButton("Open Folder", QDialogButtonBox.ButtonRole.ActionRole)
        open_btn.clicked.connect(lambda: QDesktopServices.openUrl(QUrl.fromLocalFile(path_str)))
        open_btn.clicked.connect(dialog.close)
        bbox.rejected.connect(dialog.close)
        layout.addWidget(bbox)
        dialog.exec()

    def on_about(self):
        """Show the About dialog."""
        QMessageBox.about(
            self,
            "About Portfolio Tracker",
            "<h3>Portfolio Tracker</h3>"
            f"<p>Version {__version__}</p>"
            "<p>Developed by Fh666z</p>"
            "<p>Track your portfolio holdings, instrument types, and performance.</p>"
            "<p>Import from images or spreadsheets, manage currencies, and export to JSON, CSV, or Excel.</p>"
            "<p><b>License:</b> MIT License</p>"
            "<p><b>Prerequisites:</b> Tesseract OCR is optional for image import. "
            "<a href='https://github.com/UB-Mannheim/tesseract/wiki'>Install Tesseract (Windows)</a></p>"
        )
    
    def setup_shortcuts(self):
        """Set up keyboard shortcuts for common operations."""
        # Ctrl+N: New Input
        shortcut_new = QShortcut(QKeySequence("Ctrl+N"), self)
        shortcut_new.activated.connect(self.on_new_input)
        
        # Ctrl+S: Save Data
        shortcut_save = QShortcut(QKeySequence("Ctrl+S"), self)
        shortcut_save.activated.connect(self.on_save_data)
        
        # Ctrl+O: Load/Open Data
        shortcut_load = QShortcut(QKeySequence("Ctrl+O"), self)
        shortcut_load.activated.connect(self.on_load_data)
        
        # Ctrl+F: Focus search box (if on Portfolio tab)
        shortcut_search = QShortcut(QKeySequence("Ctrl+F"), self)
        shortcut_search.activated.connect(self.focus_search)
    
    def focus_search(self):
        """Focus the search box in the Portfolio tab."""
        # Switch to Portfolio tab if not already there
        portfolio_index = self.tabs.indexOf(self.portfolio_tab)
        if portfolio_index >= 0:
            self.tabs.setCurrentIndex(portfolio_index)
        # Focus the search input
        self.portfolio_tab.search_input.setFocus()
        self.portfolio_tab.search_input.selectAll()
    
    def closeEvent(self, event):
        """Handle window close event."""
        # Save window geometry
        self.settings_store.set('window_geometry', self.saveGeometry().toHex().data().decode())
        
        # Save portfolio and mappings (no feedback needed when closing)
        self.save_all(show_feedback=False)
        
        event.accept()
    
    def refresh_all(self):
        """Refresh all views."""
        self.portfolio_tab.refresh()
        self.config_tab.refresh()
        self.stats_tab.refresh()
        self.update_status_bar()
    
    def update_status_bar(self):
        """Update status bar with portfolio info."""
        summary = self.calculator.get_summary()
        self.status_bar.showMessage(
            f"Holdings: {summary['num_holdings']} | "
            f"Total Invested (EUR): €{summary['total_invested_eur']:,.2f} | "
            f"Total (EUR): €{summary['total_eur']:,.2f}"
        )
    
    def save_all(self, show_feedback: bool = True):
        """Save all data.
        
        Args:
            show_feedback: Whether to show "Changes saved" feedback in status bar
        """
        # Save portfolio
        self.portfolio_store.save(self.calculator.portfolio)
        
        # Save mappings from current holdings
        self.mappings_store.update_from_holdings(self.calculator.portfolio.holdings)
        
        # Save free cash to settings
        self.settings_store.set('free_cash', self.calculator.portfolio.free_cash)
        
        # Show brief feedback and then restore normal status
        if show_feedback:
            self.show_save_feedback()
    
    def show_save_feedback(self):
        """Show brief save feedback, then restore normal status bar."""
        # Show "Changes saved" briefly
        summary = self.calculator.get_summary()
        self.status_bar.showMessage(
            f"Changes saved  |  Holdings: {summary['num_holdings']} | "
            f"Total (EUR): €{summary['total_eur']:,.2f}"
        )
        # Restore normal status bar after 2 seconds
        QTimer.singleShot(2000, self.update_status_bar)
    
    def on_portfolio_changed(self):
        """Handle portfolio data change."""
        self.config_tab.refresh()
        self.stats_tab.refresh()
        self.update_status_bar()
        self.save_all()
    
    def on_config_changed(self):
        """Handle instrument configuration change."""
        self.portfolio_tab.refresh()
        self.stats_tab.refresh()
        self.update_status_bar()
        self.save_all()
    
    def on_rates_changed(self):
        """Handle exchange rate change."""
        # Refresh all views to recalculate with new rates
        self.refresh_all()
    
    def on_tab_moved(self, from_index: int, to_index: int):
        """Handle tab reorder - save new order."""
        # Get current tab order by tab names
        order = []
        for i in range(self.tabs.count()):
            widget = self.tabs.widget(i)
            # Find the tab name based on widget
            if widget == self.portfolio_tab:
                order.append("portfolio")
            elif widget == self.config_tab:
                order.append("instrument_config")
            elif widget == self.stats_tab:
                order.append("statistics")
            elif widget == self.currency_tab:
                order.append("currency_exchange")
        self.settings_store.set_tab_order(order)
    
    def restore_tab_order(self):
        """Restore saved tab order."""
        order = self.settings_store.get_tab_order()
        if not order:
            return
        
        # Map tab names to widgets
        name_to_widget = {
            "portfolio": self.portfolio_tab,
            "instrument_config": self.config_tab,
            "statistics": self.stats_tab,
            "currency_exchange": self.currency_tab,
        }
        
        # Reorder tabs based on saved order
        for target_index, name in enumerate(order):
            widget = name_to_widget.get(name)
            if widget:
                current_index = self.tabs.indexOf(widget)
                if current_index != -1 and current_index != target_index:
                    self.tabs.tabBar().moveTab(current_index, target_index)
    
    def on_new_input(self):
        """Handle new input button click."""
        dialog = ImportDialog(self)
        dialog.file_selected.connect(self.process_import_file)
        dialog.exec()
    
    def on_reset_data(self):
        """Handle reset data button click."""
        # First warning dialog
        reply = QMessageBox.warning(
            self,
            "Reset Portfolio Data",
            "This will erase ALL portfolio data including:\n"
            "  - All holdings\n"
            "  - Free cash balance\n\n"
            "This action cannot be undone.\n\n"
            "Are you sure you want to continue?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # Second confirmation: require typing "DELETE"
        text, ok = QInputDialog.getText(
            self,
            "Confirm Reset",
            "Type DELETE to confirm:",
        )
        
        if ok and text.strip().upper() == "DELETE":
            # Clear all holdings
            self.calculator.portfolio.holdings.clear()
            # Reset free cash
            self.calculator.set_free_cash(0)
            # Refresh all views
            self.refresh_all()
            # Save the empty state
            self.save_all()
            # Show confirmation
            self.status_bar.showMessage("Portfolio data has been reset.", 5000)
        elif ok:
            QMessageBox.information(
                self,
                "Reset Cancelled",
                "Reset was cancelled. You must type 'DELETE' to confirm."
            )
    
    def on_save_data(self):
        """Handle save data button click."""
        # Generate default filename with current date
        today = datetime.now()
        default_filename = f"allocation_data_{today.strftime('%Y.%m.%d')}.json"
        
        # Open save file dialog
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Portfolio Data",
            default_filename,
            "JSON Files (*.json);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            # Collect all data
            export_data = {
                "version": "1.0",
                "exported_at": datetime.now().isoformat(),
                "portfolio": {
                    "holdings": [h.to_dict() for h in self.calculator.portfolio.holdings],
                    "free_cash": self.calculator.portfolio.free_cash
                },
                "settings": {
                    "currencies": self.settings_store.get_currencies(),
                    "exchange_rates": self.settings_store.get_exchange_rates()
                },
                "mappings": self.mappings_store.mappings
            }
            
            # Write to file
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2)
            
            self.status_bar.showMessage(f"Data saved to {file_path}", 5000)
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Save Error",
                f"Error saving data:\n{str(e)}"
            )
    
    def on_export_csv(self):
        """Export portfolio data to CSV format."""
        today = datetime.now()
        default_filename = f"portfolio_{today.strftime('%Y.%m.%d')}.csv"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to CSV",
            default_filename,
            "CSV Files (*.csv);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            holdings = self.calculator.portfolio.holdings
            
            if not holdings:
                QMessageBox.warning(self, "No Data", "No holdings to export.")
                return
            
            # Define CSV columns
            columns = [
                'Instrument', 'Position', 'Last Price', 'Currency',
                'Market Value', 'Market Value (EUR)', 'Cost Basis',
                'Target %', 'Allocation %', 'Asset Type', 'Region',
                'Unrealized P&L', 'Daily P&L'
            ]
            
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(columns)
                
                allocations = self.calculator.get_allocations()
                alloc_map = {a.instrument: a for a in allocations}
                
                for holding in holdings:
                    alloc = alloc_map.get(holding.instrument)
                    market_value_eur = self.settings_store.convert_to_eur(
                        holding.market_value, holding.currency
                    )
                    
                    row = [
                        holding.instrument,
                        holding.position,
                        holding.last_price,
                        holding.currency,
                        holding.market_value,
                        market_value_eur,
                        holding.cost_basis,
                        holding.target_allocation * 100,
                        (alloc.allocation_with_cash * 100) if alloc else 0,
                        holding.asset_type.value,
                        holding.region.value,
                        holding.unrealized_pnl,
                        holding.daily_pnl
                    ]
                    writer.writerow(row)
            
            self.status_bar.showMessage(f"Exported to {file_path}", 5000)
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Error",
                f"Error exporting to CSV:\n{str(e)}"
            )
    
    def on_export_excel(self):
        """Export portfolio data to Excel format."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            QMessageBox.warning(
                self,
                "Missing Dependency",
                "openpyxl is required for Excel export.\n"
                "Install it with: pip install openpyxl"
            )
            return
        
        today = datetime.now()
        default_filename = f"portfolio_{today.strftime('%Y.%m.%d')}.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to Excel",
            default_filename,
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            holdings = self.calculator.portfolio.holdings
            
            if not holdings:
                QMessageBox.warning(self, "No Data", "No holdings to export.")
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Portfolio"
            
            # Define columns
            columns = [
                'Instrument', 'Position', 'Last Price', 'Currency',
                'Market Value', 'Market Value (EUR)', 'Cost Basis',
                'Target %', 'Allocation %', 'Asset Type', 'Region',
                'Unrealized P&L', 'Daily P&L'
            ]
            
            # Header styling
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font_white = Font(bold=True, color='FFFFFF')
            
            # Write headers
            for col, header in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            allocations = self.calculator.get_allocations()
            alloc_map = {a.instrument: a for a in allocations}
            
            for row_idx, holding in enumerate(holdings, 2):
                alloc = alloc_map.get(holding.instrument)
                market_value_eur = self.settings_store.convert_to_eur(
                    holding.market_value, holding.currency
                )
                
                data = [
                    holding.instrument,
                    holding.position,
                    holding.last_price,
                    holding.currency,
                    holding.market_value,
                    market_value_eur,
                    holding.cost_basis,
                    holding.target_allocation * 100,
                    (alloc.allocation_with_cash * 100) if alloc else 0,
                    holding.asset_type.value,
                    holding.region.value,
                    holding.unrealized_pnl,
                    holding.daily_pnl
                ]
                
                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # Right-align numeric columns
                    if col_idx > 1:
                        cell.alignment = Alignment(horizontal='right')
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 30)
            
            # Add summary row
            summary_row = len(holdings) + 3
            ws.cell(row=summary_row, column=1, value="Total Holdings:").font = Font(bold=True)
            ws.cell(row=summary_row, column=2, value=len(holdings))
            
            summary = self.calculator.get_summary()
            ws.cell(row=summary_row + 1, column=1, value="Total Value (EUR):").font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=2, value=f"€{summary['total_eur']:,.2f}")
            
            wb.save(file_path)
            self.status_bar.showMessage(f"Exported to {file_path}", 5000)
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Error",
                f"Error exporting to Excel:\n{str(e)}"
            )
    
    def on_load_data(self):
        """Handle load data button click."""
        # Open file dialog
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Load Portfolio Data",
            "",
            "JSON Files (*.json);;All Files (*)"
        )
        
        if not file_path:
            return
        
        self.status_bar.showMessage("Loading data...")
        QApplication.processEvents()
        
        try:
            # Read and parse file
            with open(file_path, 'r', encoding='utf-8') as f:
                import_data = json.load(f)
            
            # Validate version
            version = import_data.get("version", "1.0")
            if version != "1.0":
                QMessageBox.warning(
                    self,
                    "Version Warning",
                    f"File version {version} may not be fully compatible."
                )
            
            # Load portfolio data
            portfolio_data = import_data.get("portfolio", {})
            holdings = [Holding.from_dict(h) for h in portfolio_data.get("holdings", [])]
            free_cash = float(portfolio_data.get("free_cash", 0))
            
            # Load settings
            settings_data = import_data.get("settings", {})
            if "currencies" in settings_data:
                self.settings_store.set_currencies(settings_data["currencies"])
            if "exchange_rates" in settings_data:
                for currency, rate in settings_data["exchange_rates"].items():
                    self.settings_store.set_exchange_rate(currency, rate)
            
            # Load mappings
            mappings_data = import_data.get("mappings", {})
            if mappings_data:
                self.mappings_store.mappings = mappings_data
                self.mappings_store.save()
            
            # Apply to portfolio
            self.calculator.portfolio.holdings = holdings
            self.calculator.set_free_cash(free_cash)
            
            # Apply mappings to loaded holdings
            self.mappings_store.apply_mappings(self.calculator.portfolio.holdings)
            
            # Refresh all views
            self.refresh_all()
            
            # Save to internal storage
            self.save_all()
            
            self.status_bar.showMessage(
                f"Loaded {len(holdings)} holdings from {file_path}", 5000
            )
            
        except json.JSONDecodeError as e:
            QMessageBox.critical(
                self,
                "Load Error",
                f"Invalid JSON file:\n{str(e)}"
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Load Error",
                f"Error loading data:\n{str(e)}"
            )
    
    def process_import_file(self, file_path: str):
        """Process an imported file."""
        path = Path(file_path)
        suffix = path.suffix.lower()
        
        # Check Tesseract availability for images before showing progress
        if suffix in ('.png', '.jpg', '.jpeg'):
            if not check_tesseract():
                QMessageBox.warning(
                    self,
                    "Tesseract Not Found",
                    "Tesseract OCR is required for image processing.\n\n"
                    "Please install Tesseract:\n"
                    "- Windows: Download from https://github.com/UB-Mannheim/tesseract/wiki\n"
                    "- Make sure to add it to your PATH"
                )
                return
        
        # Show progress dialog
        is_image = suffix in ('.png', '.jpg', '.jpeg')
        progress_text = "Processing image with OCR..." if is_image else "Importing file..."
        
        progress = QProgressDialog(progress_text, None, 0, 0, self)
        progress.setWindowTitle("Importing")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)  # Show immediately
        progress.setValue(0)
        progress.show()
        QApplication.processEvents()  # Ensure dialog is displayed
        
        try:
            # Update status bar
            self.status_bar.showMessage(progress_text)
            
            # Parse file based on type
            if is_image:
                holdings = parse_image_file(file_path)
            else:
                holdings = parse_file(file_path)
            
            progress.close()
            
            if not holdings:
                QMessageBox.warning(
                    self,
                    "No Data Found",
                    f"Could not extract any portfolio data from:\n{file_path}"
                )
                self.status_bar.showMessage("Import failed: no data found", 3000)
                return
            
            self.status_bar.showMessage(f"Found {len(holdings)} holdings", 2000)
            
            # Show review dialog
            review_dialog = ReviewDialog(holdings, file_path, self)
            review_dialog.data_confirmed.connect(self.on_data_confirmed)
            review_dialog.exec()
            
        except Exception as e:
            progress.close()
            self.status_bar.showMessage("Import failed", 3000)
            QMessageBox.critical(
                self,
                "Import Error",
                f"Error importing file:\n{str(e)}"
            )
    
    def on_data_confirmed(self, holdings: list[Holding]):
        """Handle confirmed import data."""
        # Apply existing mappings to new holdings
        self.mappings_store.apply_mappings(holdings)
        
        # Add/update holdings in portfolio
        self.calculator.portfolio.add_or_update_holdings(holdings)
        
        # Refresh views
        self.refresh_all()
        
        # Save
        self.save_all()
        
        # Show success message
        self.status_bar.showMessage(
            f"Imported {len(holdings)} holdings successfully!",
            5000  # Show for 5 seconds
        )
