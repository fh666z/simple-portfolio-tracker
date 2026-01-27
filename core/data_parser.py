"""Excel/CSV data parser for Portfolio Tracker."""
import re
from pathlib import Path
from typing import Optional
from openpyxl import load_workbook

from .models import Holding


def parse_number(value: any) -> float:
    """Parse a number from various formats."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    
    # Handle string values
    s = str(value).strip()
    
    # Remove currency prefix (e.g., 'C31.86' for CNY)
    if s.startswith('C') and len(s) > 1:
        s = s[1:]
    
    # Handle dash/em-dash for missing values
    if s in ('—', '-', '--', ''):
        return 0.0
    
    # Remove commas and parse
    s = s.replace(',', '')
    
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_percentage(value: any) -> float:
    """Parse a percentage value (can be decimal like 0.05 or string like '5%')."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        # Assume it's already in decimal form if small, otherwise percentage
        return float(value)
    
    s = str(value).strip()
    
    # Handle dash for missing values
    if s in ('—', '-', '--', ''):
        return 0.0
    
    # Remove % sign if present
    if s.endswith('%'):
        s = s[:-1]
        return float(s) / 100
    
    try:
        return float(s)
    except ValueError:
        return 0.0


def clean_instrument_name(name: str) -> str:
    """Clean instrument name (remove trailing spaces, non-breaking spaces)."""
    if name is None:
        return ""
    return str(name).strip().replace('\xa0', '').strip()


def parse_excel_file(file_path: str | Path) -> list[Holding]:
    """
    Parse an Excel file and return list of Holdings.
    
    Expected columns (flexible order):
    - Instrument
    - Position
    - Last (price)
    - Change %
    - Cost Basis
    - Market Value
    - Avg Price
    - Daily P&L
    - Unrealized P&L
    """
    file_path = Path(file_path)
    
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # Find header row and column mapping
    header_row = None
    column_map = {}
    
    # Column name variations to handle
    # Note: Order matters - more specific matches should come first to avoid
    # substring matching issues (e.g., 'price' matching 'avg price')
    column_aliases = {
        'instrument': ['instrument', 'ticker', 'symbol', 'name'],
        'position': ['position', 'qty', 'quantity', 'shares', 'units'],
        'avg_price': ['avg price', 'average price', 'avgprice'],  # Check before last_price
        'last_price': ['last', 'last price', 'current price', 'lastprice'],  # Removed generic 'price'
        'change_pct': ['change %', 'change', 'chg %', 'daily change'],
        'cost_basis': ['cost basis', 'cost', 'total cost', 'basis'],
        'market_value': ['market value', 'value', 'mkt value', 'current value'],
        'daily_pnl': ['daily p&l', 'daily pnl', 'day p&l', 'daily gain'],
        'unrealized_pnl': ['unrealized p&l', 'unrealized pnl', 'unrealized', 'total p&l', 'gain/loss'],
    }
    
    # Search for header row (first 10 rows)
    for row_idx, row in enumerate(ws.iter_rows(max_row=10), 1):
        row_values = [str(cell.value).lower().strip() if cell.value else '' for cell in row]
        
        # Check if this looks like a header row
        matches = 0
        temp_column_map = {}
        
        for col_idx, val in enumerate(row_values):
            val_clean = val.replace('\xa0', ' ').strip()
            
            # Find the best matching field (prefer longer/more specific matches)
            best_match = None
            best_match_len = 0
            
            for field, aliases in column_aliases.items():
                if field in temp_column_map:
                    continue  # Already mapped this field
                for alias in aliases:
                    if alias in val_clean and len(alias) > best_match_len:
                        best_match = field
                        best_match_len = len(alias)
            
            if best_match:
                temp_column_map[best_match] = col_idx
                matches += 1
        
        if matches >= 3:  # At least 3 recognized columns
            column_map = temp_column_map
            header_row = row_idx
            break
    
    if header_row is None:
        raise ValueError("Could not find header row in Excel file")
    
    # Parse data rows
    holdings = []
    for row in ws.iter_rows(min_row=header_row + 1):
        # Get instrument name
        instrument_col = column_map.get('instrument', 0)
        instrument = clean_instrument_name(row[instrument_col].value)
        
        # Skip empty rows
        if not instrument:
            continue
        
        # Skip summary rows
        if any(keyword in instrument.lower() for keyword in ['total', 'sum', 'pending']):
            continue
        
        try:
            holding = Holding(
                instrument=instrument,
                position=parse_number(row[column_map.get('position', 1)].value),
                last_price=parse_number(row[column_map.get('last_price', 2)].value),
                change_pct=parse_percentage(row[column_map.get('change_pct', 3)].value),
                cost_basis=parse_number(row[column_map.get('cost_basis', 4)].value),
                market_value=parse_number(row[column_map.get('market_value', 5)].value),
                avg_price=parse_number(row[column_map.get('avg_price', 6)].value),
                daily_pnl=parse_number(row[column_map.get('daily_pnl', 7)].value),
                unrealized_pnl=parse_number(row[column_map.get('unrealized_pnl', 8)].value),
            )
            holdings.append(holding)
        except Exception as e:
            # Skip rows that can't be parsed
            print(f"Warning: Could not parse row for {instrument}: {e}")
            continue
    
    wb.close()
    return holdings


def parse_csv_file(file_path: str | Path) -> list[Holding]:
    """Parse a CSV file and return list of Holdings."""
    import csv
    
    file_path = Path(file_path)
    holdings = []
    
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        
        for row in reader:
            # Normalize column names
            row_lower = {k.lower().strip(): v for k, v in row.items()}
            
            instrument = clean_instrument_name(
                row_lower.get('instrument') or row_lower.get('ticker') or row_lower.get('symbol', '')
            )
            
            if not instrument:
                continue
            
            try:
                holding = Holding(
                    instrument=instrument,
                    position=parse_number(row_lower.get('position') or row_lower.get('qty', 0)),
                    last_price=parse_number(row_lower.get('last') or row_lower.get('price', 0)),
                    change_pct=parse_percentage(row_lower.get('change %') or row_lower.get('change', 0)),
                    cost_basis=parse_number(row_lower.get('cost basis') or row_lower.get('cost', 0)),
                    market_value=parse_number(row_lower.get('market value') or row_lower.get('value', 0)),
                    avg_price=parse_number(row_lower.get('avg price') or row_lower.get('average', 0)),
                    daily_pnl=parse_number(row_lower.get('daily p&l') or row_lower.get('daily pnl', 0)),
                    unrealized_pnl=parse_number(row_lower.get('unrealized p&l') or row_lower.get('unrealized', 0)),
                )
                holdings.append(holding)
            except Exception as e:
                print(f"Warning: Could not parse row: {e}")
                continue
    
    return holdings


def parse_file(file_path: str | Path) -> list[Holding]:
    """Parse a file (Excel or CSV) and return list of Holdings."""
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()
    
    if suffix in ('.xlsx', '.xls'):
        return parse_excel_file(file_path)
    elif suffix == '.csv':
        return parse_csv_file(file_path)
    else:
        raise ValueError(f"Unsupported file format: {suffix}")
